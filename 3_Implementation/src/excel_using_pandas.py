# import required packages
import pandas as pd
from openpyxl import load_workbook
from matplotlib import pyplot as pt

# To read the main excel sheet
excel_file = pd.ExcelFile('book.xlsx')

# Enter the input like PS or Name or Email
something = input("ENTER PS or NAME or EMAIL as key: ")

# Checking for what input is given, if it is PS , then program flow goes to here
if something == "PS":
    PS = int(input("Enter PS: "))

# iterate for number of excel sheets inside the book.xlsx
    for i in excel_file.sheet_names:
        df1 = pd.read_excel(excel_file, i)
        # check for input
        df1.set_index('PS_Number', inplace=True)

# Lock the search for  the inputted PS
        result = df1.loc[PS]
        print(result)
    excel_file.close()

# Path for output of result into excel sheet
    path = "out.xlsx"
    book = load_workbook(path)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    if 'Mastersheet' in book.sheetnames:
        reference = book['Mastersheet']
        book.remove(reference)

# Write into excel sheet out.xlsx
    result.to_excel(writer, sheet_name='Mastersheet')

# For Plotting bar graph
    pivot = df1.groupby(['PS_Number']).mean()
    variable = pivot.loc[:, "Training_Room_5":"Team_No_5"]
    variable.plot(kind='bar')

# Show th bar graph in the window
    pt.show()
    writer.save()
    writer.close()

elif something == "NAME":

    # Checking for what input is given, if it is PS , then program flow goes to here
    NAME = input("Enter NAME: ")

    # iterate for number of excel sheets inside the book.xlsx
    for i in excel_file.sheet_names:
        df1 = pd.read_excel(excel_file, i)
        # check for input
        df1.set_index('Name', inplace=True)

        # Lock the search for  the inputted NAME
        result = df1.loc[NAME]
        print(result)
    excel_file.close()

    # Path for output of result into excel sheet
    path = "out.xlsx"
    book = load_workbook(path)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    if 'Mastersheet' in book.sheetnames:
        reference = book['Mastersheet']
        book.remove(reference)

    # Write into excel sheet out.xlsx
    result.to_excel(writer, sheet_name='Mastersheet')

    # For Plotting bar graph
    pivot = df1.groupby(['PS_Number']).mean()
    variable = pivot.loc[:, "Training_Room_5":"Team_No_5"]
    variable.plot(kind='bar')

    # Show th bar graph in the window
    pt.show()
    writer.save()
    writer.close()


elif something == "EMAIL":

    # Checking for what input is given, if it is PS , then program flow goes to here
    EMAIL = input("Enter Email: ")

    # iterate for number of excel sheets inside the book.xlsx
    for i in excel_file.sheet_names:
        df1 = pd.read_excel(excel_file, i)
        # check for input
        df1.set_index('Email_Address', inplace=True)

        # Lock the search for  the inputted EMAIL
        result = df1.loc[EMAIL]
        print(result)
    excel_file.close()

    # Path for output of result into excel sheet
    path = "out.xlsx"
    book = load_workbook(path)

    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    if 'Mastersheet' in book.sheetnames:
        reference = book['Mastersheet']
        book.remove(reference)

    # Write into excel sheet out.xlsx
    result.to_excel(writer, sheet_name='Mastersheet')

    # For Plotting bar graph
    pivot = df1.groupby(['PS_Number']).mean()
    variable = pivot.loc[:, "Training_Room_5":"Team_No_5"]
    variable.plot(kind='bar')

    # Show the bar graph in the window
    pt.show()
    writer.save()
    writer.close()

else:
    print("Valid keyword not entered Dude!")

excel_file.close()
